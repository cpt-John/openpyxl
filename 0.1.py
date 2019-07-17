import openpyxl as xl
# pip install openpyxl
wb = xl.load_workbook('xls_data.xlsx')
sheet = wb.worksheets[0]


dict_o = {'N': 0, 'R': 1, 'S': 2, 'O': 3, 'A': 4}
dict_r = {'N': 4, 'R': 3, 'S': 2, 'O': 1, 'A': 0}


def replace_col(col_id, dict_tyf):
    i = 2
    while i < 88:  # rows
        cell = sheet.cell(i, col_id)
        cell.value = dict_tyf.get(cell.value[0], '!')
        i += 1


c = 18  # columns start
list_or = [18, 20, 22, 23, 25, 26, 30, 33]  # correct order
while c < 34:  # columns end
    if c in list_or:
        dict_ty = dict_o
    else:
        dict_ty = dict_r
    replace_col(c, dict_ty)
    c += 1

print(sheet.cell(1, 1))
wb.save('test.xlsx')
